#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Screw Length Extraction – Version 2 (Hybrid B+C Approach)

Improvements over v1:
- Two-stage processing: raw extraction → business rules application
- Enhanced diagnostics and logging
- Separate sheet for skipped edge cases
- More flexible regex patterns
- Better handling of cut instructions and weighted averages
- Priority system: item code suffix > description > notes (excluding parentheses)

Author: Enhanced version with hybrid approach
"""

import argparse
import json
import math
import re
import time
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Union
from enum import Enum

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------------------
# Config & constants
# ---------------------------

DEFAULT_LEN_COL = "OFFICIAL LENGTH (INCHES)"
DEFAULT_QTY_COL = "OFFICIAL QUANTITY (PCS.)"

EACHLIKE_UOMS = {"EA", "EACH", "PCS", "PC", "PIECE", "PIECES"}
FEETLIKE_UOMS = {"FT", "FEET", "FOOT"}

# Non-screw items to skip
NON_SCREW_KEYWORDS = [
    "FLANGE", "BRACKET", "WASHER", "BEARING", "BUSHING", "COLLAR", 
    "COUPLING", "MOUNT", "ADAPTER", "FITTING", "SEAL", "GASKET",
    "O-RING", "SPRING", "PIN", "KEY", "SHIM", "SPACER", "PLATE"
]

# Valid screw-related keywords - expanded
SCREW_KEYWORDS = ["SCREW", "WORM", "KNURL", "ACME", "THREAD", "STUB", "BOLT", "STUD", "FASTENER", "RIVET", "DOWEL", "ROD"]

# Contextual screw phrases that should override non-screw classification
CONTEXTUAL_SCREW_PHRASES = [
    "MOUNT SCREW", "MOUNTING SCREW", "BRACKET SCREW", "ADAPTER SCREW",
    "COUPLING SCREW", "FLANGE SCREW", "THREADED ROD", "THREADED STUD",
    "SHOULDER SCREW", "MACHINE SCREW", "HEX BOLT", "CAP SCREW"
]

class SkipReason(Enum):
    """Reasons for skipping a record"""
    NON_SCREW_ITEM = "Non-screw item detected"
    RL_COMPLEX = "Complex RL calculation without clear piece count"
    NO_LENGTH_AVAILABLE = "No length information found"
    PARSE_ERROR = "Unable to parse length/quantity"
    OVER_MAX_LENGTH = "Length exceeds 300 inches"

@dataclass
class ExtractConfig:
    max_inch: float = 300.0
    prefer_inches_in_parentheses: bool = True  # Enable parentheses preference
    rl_default_bar_feet: float = 12.0
    ft_tolerance_pct: float = 0.05  # 5% tolerance (tightened from 10%)
    round_length_ndp: int = 2       # round to 2 decimals
    use_round_for_ft_qty: bool = True
    enable_diagnostics: bool = True
    debug_mode: bool = False        # Enable comprehensive diagnostic logging
    rl_fallback_enabled: bool = False  # Enable RL fallback for zero-info cases
    # Column names
    col_item_code: str = "ITEM CODE"
    col_desc: str = "ITEM DESCRIPTION"
    col_notes: str = "NOTE"
    col_orig_qty: str = "ORIGINAL QUANTITY SHIP/REC"
    col_orig_uom: str = "ORIGINAL UNIT OF MEASURE"
    col_official_len: str = DEFAULT_LEN_COL
    col_official_qty: str = DEFAULT_QTY_COL

# ---------------------------
# Enhanced regex patterns (more flexible)
# ---------------------------

# Item code patterns - Updated to support decimals
RE_CODE_SUFFIX_NUM = re.compile(r"-(\d+(?:\.\d+)?)$")
RE_CODE_SUFFIX_L = re.compile(r"-L$", re.IGNORECASE)

# Parentheses patterns (for reference but lower priority)
RE_INCH_PAREN = re.compile(r"\((\s*\d+(?:\.\d+)?\s*)(?:\"|\s*in(?:ch(?:es)?)?)\s*\)", re.IGNORECASE)
RE_MM_PAREN = re.compile(r"\((\s*\d+(?:\.\d+)?)\s*mm\s*\)", re.IGNORECASE)

# Length patterns
RE_FT_IN = re.compile(
    r"(?P<ft>\d+)\s*['']\s*(?P<in>(?:\d+(?:\.\d+)?|\d+\s*-\s*\d+/\d+|\d+/\d+)?)\s*(?:\"|'')?",
    re.IGNORECASE,
)
RE_IN_ONLY = re.compile(
    r"(?P<in>\d+(?:\.\d+)?|\d+\s*-\s*\d+/\d+|\d+/\d+)\s*(?:\"|''|\bIN\b|\bINCH(?:ES)?\b)",
    re.IGNORECASE,
)
RE_MM_ONLY = re.compile(
    r"(?P<mm>\d+(?:\.\d+)?)\s*MM\b",
    re.IGNORECASE,
)
RE_CM_ONLY = re.compile(
    r"(?P<cm>\d+(?:\.\d+)?)\s*CM\b",
    re.IGNORECASE,
)

# Metric thread pattern for implicit metric detection
RE_METRIC_THREAD = re.compile(r"M\d+", re.IGNORECASE)

# X-by pattern
RE_X_BY = re.compile(r"\b(?:x|×|by)\b", re.IGNORECASE)

# Enhanced cut plan patterns - much more flexible with multi-segment support
CUT_PATTERNS = [
    # Standard patterns with various formats - Enhanced for "OAL each"
    re.compile(r"(?:\*+)?\s*CUT\s+(?P<count>\d+)\s*(?:PCS?\.?|PIECES?\.?)\s*(?:TO|@)\s*(?P<len>[\d\s\-/\"'\.]+?)(?:\s*(?:OAL|OVERALL|LENGTH)(?:\s+EACH|EA\.?)?|\s*(?:EACH|EA\.?))?\s*(?:\*+)?", re.IGNORECASE),
    # Reversed order: "CUT TO X" THEN "Y PCS" - Enhanced for "OAL each"  
    re.compile(r"(?:\*+)?\s*CUT\s+TO\s+(?P<len>[\d\s\-/\"'\.]+?)\s*(?:OAL(?:\s+EACH)?)??\s*(?P<count>\d+)\s*(?:PCS?\.?|PIECES?\.?)", re.IGNORECASE),
    # Simple "X PCS TO Y" or "X PCS @ Y" - Enhanced for "OAL each"
    re.compile(r"(?P<count>\d+)\s*(?:PCS?\.?|PIECES?\.?|PC\.?)\s*(?:TO|@)\s*(?P<len>[\d\s\-/\"'\.]+?)(?:\s*(?:OAL|LENGTH|LENGTHS)(?:\s+EACH)?|\s*EACH)?", re.IGNORECASE),
    # "X PIECES OF Y LENGTH"
    re.compile(r"(?P<count>\d+)\s*(?:PCS?\.?|PIECES?\.?)\s*(?:OF|AT)\s*(?P<len>[\d\s\-/\"'\.]+?)\s*(?:LENGTH|LONG)?", re.IGNORECASE),
    # NEW: Standalone piece declarations like "1 PC 3'" or "2 PCS 24""
    re.compile(r"(?P<count>\d+)\s*(?:PC\.?|PCS?\.?|PIECES?\.?)\s+(?P<len>[\d\s\-/\"'\.]+?)(?:\s|$|;|AND|&)", re.IGNORECASE),
    # NEW: Piece declarations with "of" like "2 pieces of 6'"
    re.compile(r"(?P<count>\d+)\s*(?:PIECE?S?)\s*OF\s*(?P<len>[\d\s\-/\"'\.]+?)(?:\s|$|;|AND|&)", re.IGNORECASE),
    # NEW: Length-first format like "6' x 2 pieces" or "24\" - 3 pcs"
    re.compile(r"(?P<len>[\d\s\-/\"'\.]+?)\s*[-x×]\s*(?P<count>\d+)\s*(?:PC\.?|PCS?\.?|PIECES?\.?)", re.IGNORECASE),
    # NEW: Quantity-length pairs like "QTY 2 @ 12'" or "2 EA @ 6'"
    re.compile(r"(?:QTY\s*)?(?P<count>\d+)\s*(?:EA\.?|EACH)?\s*@\s*(?P<len>[\d\s\-/\"'\.]+?)", re.IGNORECASE),
]

# Ambiguous cut instruction patterns with piece count inference
AMBIGUOUS_CUT_PATTERNS = [
    # Pattern with piece count inference: (regex, piece_count, description)
    (re.compile(r"CUT\s+(?:BARS?\s+)?IN\s+HALF", re.IGNORECASE), 2, "cut in half"),
    (re.compile(r"CUT\s+(?:INTO\s+)?(?:THIRDS?|1/3)", re.IGNORECASE), 3, "cut into thirds"),
    (re.compile(r"CUT\s+(?:INTO\s+)?(?:QUARTERS?|1/4)", re.IGNORECASE), 4, "cut into quarters"), 
    (re.compile(r"CUT\s+INTO\s+(\d+)\s*(?:PIECES?|PCS?)", re.IGNORECASE), None, "cut into N pieces"),  # Special case
    (re.compile(r"CUT\s+(?:BARS?\s+)?AS\s+NEEDED", re.IGNORECASE), 1, "cut as needed"),
    (re.compile(r"CUT\s+(?:BARS?\s+)?IF\s+NECESSARY", re.IGNORECASE), 1, "cut if necessary"),
]

# RL patterns
RE_RL = re.compile(r"(?P<feet>\d+(?:\.\d+)?)\s*(?:'|ft|feet)\s*RL\b", re.IGNORECASE)
RE_RL_SHORT = re.compile(r"\bRL\b", re.IGNORECASE)
RE_RL_NOTE_COUNT = re.compile(r"SHIP\s+(?P<count>\d+)\s+(?:\d+['']\s*)?RL", re.IGNORECASE)

# ---------------------------
# Stage 1: Raw extraction utilities
# ---------------------------

@dataclass
class RawExtraction:
    """Results from Stage 1 raw extraction"""
    code_suffix_length: Optional[float] = None
    desc_length: Optional[float] = None
    notes_length: Optional[float] = None
    parentheses_length: Optional[float] = None
    cut_plans: List[Tuple[int, float]] = field(default_factory=list)
    is_rl: bool = False
    rl_bar_feet: Optional[float] = None
    rl_piece_count: Optional[int] = None
    diagnostics: List[str] = field(default_factory=list)

@dataclass
class ProcessedResult:
    """Results from Stage 2 processing"""
    official_length: Optional[float] = None
    official_qty: Optional[int] = None
    skip_reason: Optional[SkipReason] = None
    source: str = ""
    confidence: str = ""
    flags: List[str] = field(default_factory=list)
    rationale: str = ""
    diagnostics: List[str] = field(default_factory=list)

def _to_float_safe(x) -> Optional[float]:
    try:
        return float(x)
    except Exception:
        return None

def _strip_quotes(s: str) -> str:
    return s.replace('"', '').replace(""", "").replace(""", "").replace("'", "").replace("'", "").strip()

def _parse_fractional_number(token: str) -> Optional[float]:
    """Parse fractional/decimal numbers like 1-1/4, 1 1/4, 1/4, 1.25"""
    if token is None:
        return None
    t = _strip_quotes(token.strip().lower())
    
    # Remove trailing units if present
    t = re.sub(r'(in(?:ch(?:es)?)?|oal|overall|length|each|ea)', '', t, flags=re.IGNORECASE).strip()
    
    # whole - num/den or whole num/den
    m = re.match(r"^(\d+)\s*[-\s]\s*(\d+)\s*/\s*(\d+)$", t)
    if m:
        whole = int(m.group(1))
        num = int(m.group(2))
        den = int(m.group(3)) if int(m.group(3)) > 0 else 1
        return whole + num / den
    
    # num/den
    m = re.match(r"^(\d+)\s*/\s*(\d+)$", t)
    if m:
        num = int(m.group(1))
        den = int(m.group(2)) if int(m.group(2)) > 0 else 1
        return num / den
    
    # decimal/whole
    m = re.match(r"^(\d+(?:\.\d+)?)$", t)
    if m:
        return float(t)
    
    return None

def detect_implicit_metric(text: str, bare_number: float) -> Optional[float]:
    """Detect if a bare number is likely metric based on context"""
    if not text or bare_number <= 100:
        return None
    
    text_upper = text.upper()
    
    # If text contains metric thread designation and number is large, likely mm
    if RE_METRIC_THREAD.search(text_upper) and bare_number > 100:
        return bare_number / 25.4  # Convert mm to inches
    
    # If description mentions metric terms and number is very large, likely mm
    metric_indicators = ["METRIC", "ISO", "DIN", "M10", "M12", "M16", "M20"]
    if any(indicator in text_upper for indicator in metric_indicators) and bare_number > 300:
        return bare_number / 25.4
    
    return None

def parse_len_token_to_inches(token: str, context_text: str = "") -> Optional[float]:
    """Convert a length token to inches with enhanced metric support"""
    if token is None:
        return None
    s = token.strip()
    
    # Feet-inches (e.g., 4' 6", 6')
    m = RE_FT_IN.search(s)
    if m:
        ft = int(m.group("ft"))
        in_part = m.group("in") if m.group("in") else "0"
        in_val = _parse_fractional_number(in_part)
        if in_val is None:
            in_val = 0
        return ft * 12.0 + in_val
    
    # Inches only
    m = RE_IN_ONLY.search(s)
    if m:
        return _parse_fractional_number(m.group("in"))
    
    # CM (centimeters)
    m = RE_CM_ONLY.search(s)
    if m:
        cm = _to_float_safe(m.group("cm"))
        if cm is not None:
            return (cm * 10) / 25.4  # Convert cm to mm, then to inches
    
    # MM only
    m = RE_MM_ONLY.search(s)
    if m:
        mm = _to_float_safe(m.group("mm"))
        if mm is not None:
            return mm / 25.4
    
    # Try implicit metric detection for bare numbers
    bare_num = _to_float_safe(s)
    if bare_num is not None and context_text:
        implicit_inches = detect_implicit_metric(context_text, bare_num)
        if implicit_inches is not None:
            return implicit_inches
    
    return None

def extract_inches_from_parentheses(text: str) -> Optional[float]:
    """Extract inches from parentheses - lower priority per requirements"""
    if not text:
        return None
    
    # Look for explicit inches in parentheses
    m = RE_INCH_PAREN.search(text)
    if m:
        return _to_float_safe(m.group(1))
    
    # Look for mm in parentheses (but we prefer inches if both present)
    m = RE_MM_PAREN.search(text)
    if m:
        mm = _to_float_safe(m.group(1))
        if mm is not None:
            return mm / 25.4
    
    return None

def extract_len_from_xby(text: str) -> Optional[float]:
    """Extract length from x-by-y patterns"""
    if not text:
        return None
    parts = RE_X_BY.split(text)
    if len(parts) < 2:
        return None
    tail = parts[-1]
    return parse_len_token_to_inches(tail, text)

def find_all_cut_plans(text: str, diagnostics: List[str]) -> List[Tuple[int, float]]:
    """Find all cut instructions in text using flexible patterns with multi-segment support"""
    if not text:
        return []
    
    out: List[Tuple[int, float]] = []
    text_upper = text.upper()
    
    # Split text on common delimiters for multi-segment parsing
    segments = []
    # Split on semicolons, "and", "&", or multiple spaces
    import re as re_module
    segment_splitter = re_module.compile(r'[;&]|\s+AND\s+|\s+&\s+|\s{3,}', re_module.IGNORECASE)
    segments = [seg.strip() for seg in segment_splitter.split(text) if seg.strip()]
    
    # If no splitting occurred, use original text
    if len(segments) <= 1:
        segments = [text]
    
    # Process each segment
    for seg_idx, segment in enumerate(segments):
        segment_matches = []
        
        # Try each pattern on this segment
        for pattern_idx, pattern in enumerate(CUT_PATTERNS):
            for m in pattern.finditer(segment):
                try:
                    cnt = int(m.group("count"))
                    len_str = m.group("len")
                    inches = parse_len_token_to_inches(len_str, segment)
                    if inches is not None and inches > 0:
                        segment_matches.append((cnt, inches))
                        diagnostics.append(f"Cut pattern {pattern_idx+1} matched in segment {seg_idx+1}: {cnt} pcs @ {inches:.2f}\"")
                except Exception as e:
                    diagnostics.append(f"Cut pattern {pattern_idx+1} parse error in segment {seg_idx+1}: {e}")
        
        # Add segment matches to output
        out.extend(segment_matches)
        
        # Check for drop/leftover pieces mentioned in segment
        drop_pattern = re_module.compile(r'(?:DROP|LEFTOVER|REMAINDER|SCRAP)\s*(?:PIECE?S?)?\s*(?:OF\s+)?(?P<len>[\d\s\-/\"\'\.]+?)(?:\s|$)', re_module.IGNORECASE)
        for drop_match in drop_pattern.finditer(segment):
            try:
                len_str = drop_match.group("len")
                inches = parse_len_token_to_inches(len_str, segment)
                if inches is not None and inches > 0:
                    out.append((1, inches))  # Assume 1 piece for drop
                    diagnostics.append(f"Found drop piece in segment {seg_idx+1}: 1 pc @ {inches:.2f}\"")
            except Exception as e:
                diagnostics.append(f"Drop piece parse error in segment {seg_idx+1}: {e}")
    
    # Remove duplicates while preserving order
    seen = set()
    unique_out = []
    for item in out:
        if item not in seen:
            seen.add(item)
            unique_out.append(item)
    
    # Flag if complex multi-cut was processed
    if len(segments) > 1 and unique_out:
        diagnostics.append(f"Complex multi-cut instruction processed: {len(segments)} segments, {len(unique_out)} cut groups")
    
    return unique_out

def detect_ambiguous_cuts(text: str) -> bool:
    """Detect ambiguous cut instructions like 'cut in half'"""
    if not text:
        return False
    
    for pattern_tuple in AMBIGUOUS_CUT_PATTERNS:
        pattern = pattern_tuple[0]  # Extract regex from tuple
        if pattern.search(text):
            return True
    return False

def calculate_ambiguous_cut_breakdown(text: str, source_length: float) -> Optional[Tuple[int, float, str]]:
    """Calculate piece breakdown for ambiguous cut instructions
    Returns: (piece_count, length_per_piece, reasoning) or None"""
    if not text or not source_length:
        return None
    
    for pattern_tuple in AMBIGUOUS_CUT_PATTERNS:
        pattern, default_count, description = pattern_tuple
        match = pattern.search(text)
        if match:
            # Special handling for "CUT INTO N PIECES" pattern
            if description == "cut into N pieces":
                try:
                    piece_count = int(match.group(1))  # Extract N from the match
                    length_per_piece = source_length / piece_count
                    reasoning = f"calculated {piece_count} pieces from '{description}'"
                    return (piece_count, length_per_piece, reasoning)
                except (ValueError, IndexError):
                    continue
            elif default_count and default_count > 1:
                # Calculate breakdown for fixed piece counts
                length_per_piece = source_length / default_count
                reasoning = f"calculated {default_count} pieces from '{description}'"
                return (default_count, length_per_piece, reasoning)
            else:
                # For "as needed" or "if necessary", return single piece
                reasoning = f"ambiguous instruction '{description}' - using full length"
                return (1, source_length, reasoning)
    
    return None

def is_non_screw_item(desc: str) -> bool:
    """Check if item is not a screw based on description - enhanced with contextual logic"""
    if not desc:
        return False
    
    desc_upper = desc.upper()
    
    # First check for contextual screw phrases (highest priority)
    for phrase in CONTEXTUAL_SCREW_PHRASES:
        if phrase in desc_upper:
            return False  # Definitely a screw item
    
    # Check if it contains valid screw keywords
    has_screw_keyword = any(keyword in desc_upper for keyword in SCREW_KEYWORDS)
    if has_screw_keyword:
        return False  # Likely a screw item
    
    # Check for non-screw keywords only if no screw indicators found
    has_non_screw = any(keyword in desc_upper for keyword in NON_SCREW_KEYWORDS)
    return has_non_screw

def parse_item_code_suffix_inches(code: str, diagnostics: List[str] = None) -> Optional[float]:
    """Extract length from item code suffix with enhanced decimal and mm conversion support"""
    if not code:
        return None
    
    if diagnostics is None:
        diagnostics = []
    
    # Skip -L codes
    if RE_CODE_SUFFIX_L.search(code):
        return None
    
    # Look for numeric suffix (now supports decimals)
    m = RE_CODE_SUFFIX_NUM.search(code)
    if m:
        val_str = m.group(1)
        val = _to_float_safe(val_str)
        if val is None:
            return None
        
        # Check if it contains decimal (likely mm)
        if '.' in val_str:
            # Decimal values in item codes are typically mm
            val_inches = val / 25.4
            diagnostics.append(f"Item code suffix interpreted as mm: {val} -> {val_inches:.2f}\"")
            return val_inches
        # Check if it could be MM (over 300)
        elif val > 300:
            # Convert MM to inches
            val_inches = val / 25.4
            diagnostics.append(f"Item code suffix converted from mm: {val} -> {val_inches:.2f}\"")
            return val_inches
        # Check if suspiciously high for inches (but under 300)
        elif val > 100:
            # Flag as potentially suspicious but don't convert
            diagnostics.append(f"Item code suffix unusually high: {val}\" - flagging as suspect")
            return val
        else:
            return val
    
    return None

def validate_cross_sources(raw: RawExtraction) -> Tuple[bool, str]:
    """Check if multiple sources agree on length and return verification info"""
    sources = []
    
    if raw.code_suffix_length:
        sources.append(("code", raw.code_suffix_length))
    if raw.desc_length:
        sources.append(("description", raw.desc_length))
    if raw.notes_length:
        sources.append(("notes", raw.notes_length))
    
    if len(sources) < 2:
        return False, ""
    
    # Check if any two sources agree within 5%
    for i, (name1, len1) in enumerate(sources):
        for name2, len2 in sources[i+1:]:
            if abs(len1 - len2) / max(len1, len2, 0.001) <= 0.05:  # Within 5%
                return True, f"{name1} and {name2} agree ({len1:.2f}\" ≈ {len2:.2f}\")"
    
    return False, ""

def detect_explicit_units_in_notes(notes: str) -> bool:
    """Detect explicit length specifications in notes like 'Length: 24 inches'"""
    if not notes:
        return False
    
    explicit_patterns = [
        r"LENGTH:\s*\d+(?:\.\d+)?\s*(?:INCHES?|IN|\")",
        r"OVERALL\s+LENGTH:\s*\d+(?:\.\d+)?\s*(?:INCHES?|IN|\")",
        r"DIMENSION:\s*\d+(?:\.\d+)?\s*(?:INCHES?|IN|\")",
        r"SIZE:\s*\d+(?:\.\d+)?\s*(?:INCHES?|IN|\")"
    ]
    
    notes_upper = notes.upper()
    for pattern in explicit_patterns:
        if re.search(pattern, notes_upper):
            return True
    
    return False

# ---------------------------
# Stage 1: Raw Extraction
# ---------------------------

def stage1_extract_raw(row: pd.Series, cfg: ExtractConfig) -> RawExtraction:
    """Stage 1: Extract all possible length values from the row"""
    result = RawExtraction()
    
    code = str(row.get(cfg.col_item_code, "") or "")
    desc = str(row.get(cfg.col_desc, "") or "")
    notes = str(row.get(cfg.col_notes, "") or "")
    
    # 1. Item code suffix (highest priority)
    code_len = parse_item_code_suffix_inches(code, result.diagnostics)
    if code_len:
        result.code_suffix_length = code_len
        result.diagnostics.append(f"Code suffix: {code_len}\"")
    
    # 2. Check for -L codes (need to look elsewhere)
    code_is_L = bool(RE_CODE_SUFFIX_L.search(code))
    if code_is_L:
        result.diagnostics.append("Code ends with -L, checking description/notes")
    
    # 3. Extract from description
    # Try x-by pattern first
    desc_xby = extract_len_from_xby(desc)
    if desc_xby:
        result.desc_length = desc_xby
        result.diagnostics.append(f"Description x-by: {desc_xby}\"")
    else:
        # Try general parsing with context
        desc_general = parse_len_token_to_inches(desc, desc)
        if desc_general:
            result.desc_length = desc_general
            result.diagnostics.append(f"Description general: {desc_general}\"")
    
    # 4. Extract from notes (general) with context
    notes_len = parse_len_token_to_inches(notes, notes)
    if notes_len:
        result.notes_length = notes_len
        result.diagnostics.append(f"Notes general: {notes_len}\"")
    
    # 5. Extract from parentheses (lower priority)
    paren_len = extract_inches_from_parentheses(desc) or extract_inches_from_parentheses(notes)
    if paren_len:
        result.parentheses_length = paren_len
        result.diagnostics.append(f"Parentheses: {paren_len}\"")
    
    # 6. Check for cut plans
    result.cut_plans = find_all_cut_plans(notes, result.diagnostics)
    if not result.cut_plans and desc:
        # Sometimes cut plans are in description
        result.cut_plans = find_all_cut_plans(desc, result.diagnostics)
    
    # 7. Check for RL (Random Length)
    if RE_RL_SHORT.search(desc) or RE_RL_SHORT.search(notes):
        result.is_rl = True
        result.diagnostics.append("RL detected")
        
        # Try to find bar length
        m = RE_RL.search(desc) or RE_RL.search(notes)
        if m:
            result.rl_bar_feet = _to_float_safe(m.group("feet"))
        else:
            result.rl_bar_feet = cfg.rl_default_bar_feet
        
        # Try to find piece count from notes
        m = RE_RL_NOTE_COUNT.search(notes)
        if m:
            result.rl_piece_count = int(m.group("count"))
            result.diagnostics.append(f"RL piece count from notes: {result.rl_piece_count}")
    
    return result

# ---------------------------
# Stage 2: Apply Business Rules
# ---------------------------

def process_rl_item(
    row: pd.Series,
    raw: RawExtraction, 
    cfg: ExtractConfig
) -> ProcessedResult:
    """Process Random Length (RL) items with dedicated logic"""
    result = ProcessedResult()
    result.diagnostics = raw.diagnostics.copy()
    
    orig_qty = row.get(cfg.col_orig_qty, None)
    orig_uom = str(row.get(cfg.col_orig_uom, "") or "").strip().upper()
    
    # Priority 1: Use cut plans if available for RL items
    if raw.cut_plans:
        total_pcs = sum(c for c, _ in raw.cut_plans)
        total_inches = sum(c * l for c, l in raw.cut_plans)
        
        if total_pcs > 0:
            # Weighted average length for RL with cut plans
            avg_len = total_inches / total_pcs
            result.official_length = round(avg_len, 2)
            result.official_qty = total_pcs
            result.source = "rl_cut_plans"
            result.confidence = "high"
            result.rationale = f"RL with {len(raw.cut_plans)} cut group(s), weighted avg"
            result.diagnostics.append(f"RL cut plan result: {total_pcs} pcs @ {result.official_length}\"")
            result.flags.append("rl_note_inferred")
            return result
    
    # Priority 2: Use explicit piece count from SHIP X RL pattern
    if orig_uom in FEETLIKE_UOMS and orig_qty:
        total_feet = _to_float_safe(orig_qty)
        
        if raw.rl_piece_count and raw.rl_piece_count > 0:
            length_inches = (total_feet * 12.0) / raw.rl_piece_count
            result.official_length = round(length_inches, 0)  # Round to nearest inch for RL
            result.official_qty = raw.rl_piece_count
            result.source = "rl_with_count"
            result.confidence = "high"
            result.rationale = f"RL: {total_feet}ft / {raw.rl_piece_count} pcs"
            result.diagnostics.append(f"RL explicit count: {result.official_qty} pcs @ {result.official_length}\"")
            return result
    
    # Priority 3: Try to infer piece count from cut instructions in notes
    if raw.cut_plans and orig_uom in FEETLIKE_UOMS and orig_qty:
        total_feet = _to_float_safe(orig_qty)
        inferred_pieces = sum(c for c, _ in raw.cut_plans)
        
        if inferred_pieces > 0 and total_feet:
            avg_length = (total_feet * 12.0) / inferred_pieces
            result.official_length = round(avg_length, 0)  # Round to nearest inch for RL
            result.official_qty = inferred_pieces
            result.source = "rl_note_inferred"
            result.confidence = "medium"
            result.rationale = f"RL: {total_feet}ft / {inferred_pieces} pcs (inferred from notes)"
            result.diagnostics.append(f"RL inferred from notes: {result.official_qty} pcs @ {result.official_length}\"")
            result.flags.append("rl_note_inferred")
            return result
    
    # Optional RL fallback for zero-info cases
    if cfg.rl_fallback_enabled:
        # Use full RL bar length as fallback
        fallback_length = raw.rl_bar_feet * 12.0 if raw.rl_bar_feet else 144.0  # Default 12ft
        result.official_length = round(fallback_length, 0)
        result.official_qty = 1
        result.source = "rl_fallback"
        result.confidence = "low"
        result.flags.append("assumed_full_length")
        result.rationale = f"RL fallback: assumed 1 piece @ full bar length"
        result.diagnostics.append(f"RL fallback applied: 1 pc @ {result.official_length}\" (full bar)")
        return result
    
    # Skip complex RL if no piece count available and fallback disabled
    result.skip_reason = SkipReason.RL_COMPLEX
    result.diagnostics.append("Skipped: RL without clear piece count")
    return result

def process_fixed_length_item(
    row: pd.Series,
    raw: RawExtraction,
    cfg: ExtractConfig
) -> ProcessedResult:
    """Process fixed-length items with dedicated logic"""
    result = ProcessedResult()
    result.diagnostics = raw.diagnostics.copy()
    
    code = str(row.get(cfg.col_item_code, "") or "")
    notes = str(row.get(cfg.col_notes, "") or "")
    orig_qty = row.get(cfg.col_orig_qty, None)
    orig_uom = str(row.get(cfg.col_orig_uom, "") or "").strip().upper()
    
    # Priority 1: Cut plans override everything for fixed-length items
    if raw.cut_plans:
        total_pcs = sum(c for c, _ in raw.cut_plans)
        total_inches = sum(c * l for c, l in raw.cut_plans)
        
        if total_pcs > 0:
            # Weighted average length
            avg_len = total_inches / total_pcs
            result.official_length = round(avg_len, 2)
            result.official_qty = total_pcs
            result.source = "cut_plans"
            result.confidence = "high"
            result.rationale = f"{len(raw.cut_plans)} cut group(s), weighted avg"
            result.diagnostics.append(f"Cut plan result: {total_pcs} pcs @ {result.official_length}\"")
            
            # Flag complex cuts
            if len(raw.cut_plans) > 1:
                result.flags.append("complex_cut")
            return result
    
    # Determine primary source and initial confidence
    code_is_L = bool(RE_CODE_SUFFIX_L.search(code))
    
    # Priority 2: Code suffix (if not -L)
    if not code_is_L and raw.code_suffix_length:
        result.official_length = round(raw.code_suffix_length, 2)
        result.source = "item_code_suffix"
        result.confidence = "high"
        
        # Flag suspect code lengths
        if raw.code_suffix_length > 100:
            result.flags.append("suspect_code_length")
    # Priority 3: Description length
    elif raw.desc_length:
        result.official_length = round(raw.desc_length, 2)
        result.source = "description"
        result.confidence = "medium"
    # Priority 4: Notes length
    elif raw.notes_length:
        result.official_length = round(raw.notes_length, 2)
        result.source = "notes"
        result.confidence = "low"
        
        # Upgrade confidence if explicit units detected
        if detect_explicit_units_in_notes(notes):
            result.confidence = "medium"
            result.flags.append("explicit_units_in_notes")
            result.diagnostics.append("Upgraded confidence: explicit units found in notes")
    # Priority 5: Parentheses (lowest priority)
    elif raw.parentheses_length:
        result.official_length = round(raw.parentheses_length, 2)
        result.source = "parentheses"
        result.confidence = "low"
    
    # Apply cross-verification confidence boost
    if result.official_length:
        cross_verified, verification_msg = validate_cross_sources(raw)
        if cross_verified:
            if result.confidence == "medium":
                result.confidence = "high"
            elif result.confidence == "low":
                result.confidence = "medium"
            result.flags.append("cross_verified")
            result.diagnostics.append(f"Confidence boosted: {verification_msg}")
    
    # Determine quantity for fixed-length items
    if result.official_length and result.official_qty is None:
        if orig_uom in EACHLIKE_UOMS:
            # For each-like units, use original quantity
            oq = _to_float_safe(orig_qty)
            if oq is not None:
                result.official_qty = int(round(oq))
                result.diagnostics.append(f"Qty from original (EA): {result.official_qty}")
        elif orig_uom in FEETLIKE_UOMS and orig_qty:
            # For feet, calculate pieces with smart rounding
            total_feet = _to_float_safe(orig_qty)
            if total_feet and result.official_length:
                exact_qty = (total_feet * 12.0) / result.official_length
                
                # Try different rounding strategies
                qty_round = int(round(exact_qty))
                qty_floor = int(exact_qty) if exact_qty > 0 else 0
                qty_ceil = int(exact_qty) + 1 if exact_qty > int(exact_qty) else int(exact_qty)
                
                # Calculate errors for each strategy
                candidates = []
                for qty_candidate in [qty_floor, qty_round, qty_ceil]:
                    if qty_candidate > 0:
                        recon_feet = (qty_candidate * result.official_length) / 12.0
                        error = abs(recon_feet - total_feet) / max(total_feet, 0.001)
                        candidates.append((qty_candidate, error, recon_feet))
                
                # Choose the quantity with smallest error
                if candidates:
                    best_qty, best_error, best_recon = min(candidates, key=lambda x: x[1])
                    result.official_qty = best_qty
                    
                    # Check if adjustment was made
                    if best_qty != qty_round:
                        result.diagnostics.append(f"Qty adjusted to {best_qty} (from rounded {qty_round}) to fit total length better")
                    else:
                        result.diagnostics.append(f"Qty from feet calc: {result.official_qty}")
                    
                    # Flag if tolerance exceeded
                    if best_error > cfg.ft_tolerance_pct:
                        result.flags.append(f"feet_tolerance_{best_error:.1%}")
                else:
                    # Fallback to simple rounding
                    result.official_qty = int(round(exact_qty))
                    result.diagnostics.append(f"Qty from feet calc (fallback): {result.official_qty}")
    
    return result

def stage2_apply_rules(
    row: pd.Series,
    raw: RawExtraction,
    cfg: ExtractConfig
) -> ProcessedResult:
    """Stage 2: Apply business rules with modular RL vs fixed-length processing"""
    desc = str(row.get(cfg.col_desc, "") or "")
    notes = str(row.get(cfg.col_notes, "") or "")
    
    # Check if non-screw item first
    if is_non_screw_item(desc):
        result = ProcessedResult()
        result.diagnostics = raw.diagnostics.copy()
        result.skip_reason = SkipReason.NON_SCREW_ITEM
        result.diagnostics.append("Skipped: Non-screw item")
        return result
    
    # Branch into RL vs fixed-length processing
    if raw.is_rl:
        # Process RL items with specialized logic
        result = process_rl_item(row, raw, cfg)
    else:
        # Process fixed-length items with specialized logic
        result = process_fixed_length_item(row, raw, cfg)
    
    # Common post-processing for both paths
    
    # Sanity check: prefer parenthetical values over unreasonable lengths
    if (result.official_length and result.official_length > cfg.max_inch and 
        raw.parentheses_length and raw.parentheses_length <= cfg.max_inch):
        old_length = result.official_length
        result.official_length = round(raw.parentheses_length, 2)
        result.source = "parentheses_override"
        result.confidence = "medium"
        result.diagnostics.append(f"Used {result.official_length}\" from parentheses instead of {old_length}\"")
        result.flags.append("parentheses_override")
    
    # Validate length
    if result.official_length:
        if result.official_length > cfg.max_inch:
            result.skip_reason = SkipReason.OVER_MAX_LENGTH
            result.diagnostics.append(f"Skipped: Length {result.official_length} > {cfg.max_inch}")
            return result
        
        if result.official_length <= 0:
            result.official_length = None
            result.flags.append("non_positive_length")
    
    # Handle ambiguous cut instructions with intelligent interpretation
    if result.official_length is None and detect_ambiguous_cuts(notes):
        # Determine source length for calculation
        source_length = None
        source_desc = ""
        
        if raw.is_rl and raw.rl_bar_feet:
            source_length = raw.rl_bar_feet * 12.0
            source_desc = f"RL bar ({raw.rl_bar_feet}ft)"
        elif raw.code_suffix_length:
            source_length = raw.code_suffix_length
            source_desc = "item code"
        elif raw.desc_length:
            source_length = raw.desc_length
            source_desc = "description"
        elif raw.parentheses_length:
            source_length = raw.parentheses_length
            source_desc = "parentheses"
        
        if source_length:
            # Try intelligent calculation first
            breakdown = calculate_ambiguous_cut_breakdown(notes, source_length)
            if breakdown:
                piece_count, length_per_piece, reasoning = breakdown
                result.official_length = round(length_per_piece, 2)
                result.official_qty = piece_count
                result.source = "ambiguous_cut_calculated"
                result.confidence = "low"
                result.flags.append("ambiguous_cut_calculated")
                result.rationale = f"Ambiguous cut: {reasoning} from {source_desc}"
                result.diagnostics.append(f"Calculated ambiguous cut: {piece_count} pcs @ {result.official_length}\" ({reasoning})")
            else:
                # Fallback to original single-piece logic
                result.official_length = round(source_length, 2)
                result.official_qty = 1
                result.source = "ambiguous_cut_fallback"
                result.confidence = "low"
                result.flags.append("ambiguous_cut")
                result.rationale = f"Ambiguous cut instruction - using full length from {source_desc}"
                result.diagnostics.append(f"Ambiguous cut fallback: 1 pc @ {result.official_length}\" from {source_desc}")
    
    # Final validation - preserve specific skip reasons
    if result.official_length is None and result.skip_reason is None:
        result.skip_reason = SkipReason.NO_LENGTH_AVAILABLE
        result.flags.append("no_length_found")
    
    if result.official_qty is None and result.official_length is not None:
        result.flags.append("no_qty_derived")
    
    return result

# ---------------------------
# Main processing function
# ---------------------------

def process_workbook(
    input_path: str,
    sheet_name: str,
    output_path: str,
    cfg: ExtractConfig,
    add_working: bool = False
):
    """Main processing function with two-stage approach"""
    
    # Read input
    df = pd.read_excel(input_path, sheet_name=sheet_name, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    
    # Check required columns
    required = [cfg.col_item_code, cfg.col_desc, cfg.col_notes,
                cfg.col_orig_qty, cfg.col_orig_uom]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")
    
    # Process each row
    all_results = []
    skipped_rows = []
    
    for idx, row in df.iterrows():
        # Stage 1: Extract raw values
        raw = stage1_extract_raw(row, cfg)
        
        # Stage 2: Apply business rules
        processed = stage2_apply_rules(row, raw, cfg)
        
        # Store results
        all_results.append(processed)
        
        # Track skipped rows
        if processed.skip_reason:
            skipped_rows.append({
                'Row': idx + 2,  # Excel row number (1-indexed + header)
                'Item Code': row.get(cfg.col_item_code),
                'Description': row.get(cfg.col_desc),
                'Note': row.get(cfg.col_notes),
                'Skip Reason': processed.skip_reason.value,
                'Diagnostics': '; '.join(processed.diagnostics[-3:])  # Last 3 diagnostic messages
            })
    
    # Load workbook for writing
    wb = load_workbook(input_path, data_only=False, keep_vba=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    
    ws = wb[sheet_name]
    
    # Get header mapping
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            headers[str(v).strip()] = col
    
    # Ensure output columns exist
    def ensure_column(header_name):
        if header_name in headers:
            return headers[header_name]
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col).value = header_name
        headers[header_name] = new_col
        return new_col
    
    col_len_idx = ensure_column(cfg.col_official_len)
    col_qty_idx = ensure_column(cfg.col_official_qty)
    
    # Add working columns if requested
    work_cols = {}
    if add_working:
        for name in [
            "WORKING: source",
            "WORKING: confidence", 
            "WORKING: flags",
            "WORKING: rationale",
            "WORKING: diagnostics"
        ]:
            work_cols[name] = ensure_column(name)
    
    # Write results
    for i, result in enumerate(all_results, start=2):
        if not result.skip_reason:
            ws.cell(row=i, column=col_len_idx).value = result.official_length
            ws.cell(row=i, column=col_qty_idx).value = result.official_qty
        else:
            # Clear values for skipped rows
            ws.cell(row=i, column=col_len_idx).value = None
            ws.cell(row=i, column=col_qty_idx).value = None
        
        if add_working:
            ws.cell(row=i, column=work_cols["WORKING: source"]).value = result.source
            ws.cell(row=i, column=work_cols["WORKING: confidence"]).value = result.confidence
            ws.cell(row=i, column=work_cols["WORKING: flags"]).value = ", ".join(result.flags)
            ws.cell(row=i, column=work_cols["WORKING: rationale"]).value = result.rationale
            ws.cell(row=i, column=work_cols["WORKING: diagnostics"]).value = "; ".join(result.diagnostics[-3:])
    
    # Create skipped items sheet
    if skipped_rows:
        if "Skipped_Edge_Cases" in wb.sheetnames:
            del wb["Skipped_Edge_Cases"]
        
        skip_ws = wb.create_sheet("Skipped_Edge_Cases")
        skip_df = pd.DataFrame(skipped_rows)
        
        # Write headers
        for col_idx, col_name in enumerate(skip_df.columns, start=1):
            skip_ws.cell(row=1, column=col_idx).value = col_name
        
        # Write data
        for row_idx, row_data in enumerate(skip_df.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                skip_ws.cell(row=row_idx, column=col_idx).value = value
    
    # Save workbook
    wb.save(output_path)
    
    # Calculate statistics
    total_rows = len(all_results)
    skipped = len(skipped_rows)
    processed = total_rows - skipped
    with_length = sum(1 for r in all_results if r.official_length is not None)
    with_qty = sum(1 for r in all_results if r.official_qty is not None)
    
    # Confidence breakdown
    confidence_counts = {"high": 0, "medium": 0, "low": 0}
    for r in all_results:
        if r.confidence in confidence_counts:
            confidence_counts[r.confidence] += 1
    
    return {
        "total_rows": total_rows,
        "processed_rows": processed,
        "skipped_rows": skipped,
        "rows_with_length": with_length,
        "rows_with_quantity": with_qty,
        "confidence_counts": confidence_counts,
        "skipped_breakdown": {reason.value: sum(1 for r in skipped_rows if r['Skip Reason'] == reason.value) for reason in SkipReason}
    }

# ---------------------------
# Main execution
# ---------------------------

if __name__ == "__main__":
    input_file = r"C:\Users\layden\SCREWPOC\source_data.xlsx"
    output_file = r"C:\Users\layden\OneDrive - Cornerstone Solutions Group\Desktop\AI Projects\Roton\Test #4\Enhanced_Screw_Extraction_Results.xlsx"
    
    # Create config with enhanced features enabled
    cfg = ExtractConfig(
        max_inch=300.0,
        debug_mode=True,
        rl_fallback_enabled=True  # Enable RL fallback to minimize skips
    )
    
    try:
        print(f"Processing {input_file}...")
        start_time = time.time()
        
        stats = process_workbook(
            input_file,
            "Sheet1",  # Default sheet name
            output_file,
            cfg,
            add_working=True  # Add diagnostic columns
        )
        
        end_time = time.time()
        
        # Print results
        print(f"\nProcessing completed in {end_time - start_time:.2f} seconds")
        print(f"Output saved to: {output_file}")
        print(f"\nResults Summary:")
        print(f"  Total rows processed: {stats['total_rows']}")
        print(f"  Rows with extracted length: {stats['rows_with_length']}")
        print(f"  Rows with calculated quantity: {stats['rows_with_quantity']}")
        print(f"  Rows skipped: {stats['skipped_rows']}")
        
        if stats['confidence_counts']:
            print(f"\nConfidence Distribution:")
            for conf, count in stats['confidence_counts'].items():
                if count > 0:
                    print(f"  {conf.capitalize()}: {count} rows")
        
        if stats.get('skipped_breakdown'):
            print(f"\nSkipped Reasons:")
            for reason, count in stats['skipped_breakdown'].items():
                if count > 0:
                    print(f"  {reason}: {count} rows")
                    
        print(f"\nCheck the 'Skipped_Edge_Cases' sheet for details on skipped items.")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()